import pandas as pd
from openpyxl import load_workbook

# Define the range of cells for column names
start_cell = 'A8'
end_cell = 'A137'

# Create an empty dataframe to store the extracted data
dataframe = pd.DataFrame()

# List of file paths
file_paths = ['file1.xlsx', 'file2.xlsx', 'file3.xlsx']  # Replace with your file paths

for file_path in file_paths:
    # Load the workbook
    workbook = load_workbook(file_path, read_only=True)

    # Get the active sheet
    sheet = workbook.active

    # Extract the column names
    column_names = [sheet.cell(row=row, column=1).value for row in range(sheet.min_row, sheet.max_row + 1)]

    # Find the indices of the desired quarters
    q4_2020_index = column_names.index('Q4 2020')
    q4_2021_index = column_names.index('Q4 2021')
    q4_2022_index = column_names.index('Q4 2022')
    q1_2023_index = column_names.index('Q1 2023')

    # Extract the values for each quarter
    q4_2020_values = [sheet.cell(row=row, column=q4_2020_index + 1).value for row in range(sheet.min_row, sheet.max_row + 1)]
    q4_2021_values = [sheet.cell(row=row, column=q4_2021_index + 1).value for row in range(sheet.min_row, sheet.max_row + 1)]
    q4_2022_values = [sheet.cell(row=row, column=q4_2022_index + 1).value for row in range(sheet.min_row, sheet.max_row + 1)]
    q1_2023_values = [sheet.cell(row=row, column=q1_2023_index + 1).value for row in range(sheet.min_row, sheet.max_row + 1)]

    # Extract the value for the 'Credit Relationship Name' column from the second visible cell in Row 2
    credit_relationship_name = sheet.cell(row=2, column=[col for col in range(sheet.min_column, sheet.max_column + 1) if not sheet.column_dimensions[sheet.cell(row=2, column=col).column_letter].hidden][1]).value

    # Create a dataframe for the extracted data from the current file
    df = pd.DataFrame({
        'Quarter': ['Q4 2020', 'Q4 2021', 'Q4 2022', 'Q1 2023'],
        'Credit Relationship Name': [credit_relationship_name] * len(column_names),
        'Column Names': column_names,
        'Q4 2020': q4_2020_values,
        'Q4 2021': q4_2021_values,
        'Q4 2022': q4_2022_values,
        'Q1 2023': q1_2023_values
    })

    # Append the dataframe to the consolidated dataframe
    dataframe = dataframe.append(df, ignore_index=True)

# Print the consolidated dataframe
print(dataframe)
